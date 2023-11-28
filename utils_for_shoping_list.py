import openpyxl


def is_user(message):
    excel_file = openpyxl.load_workbook('users_for_shoping.xlsx')
    user_sheet = excel_file['Лист1']  # Выбираем лист в файле exel
    search = message.text


    for x in range(1, user_sheet.max_row + 1):  # Печать всех значений из столбца

        #found = user_sheet.cell(row=x, column=1).value
        if search == str(user_sheet.cell(row=x, column=1).value):
            us = user_sheet.cell(row=x, column=2).value

            return us